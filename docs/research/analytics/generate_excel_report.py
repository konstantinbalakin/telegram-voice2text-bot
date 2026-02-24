"""
Generate Excel analytics report from the Telegram Voice2Text Bot SQLite database.

Usage:
    uv run python docs/research/analytics/generate_excel_report.py

Output:
    docs/research/analytics/2026-02-24-usage-report.xlsx
"""

import sqlite3
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill

from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Config ──────────────────────────────────────────────────────────────────
DB_PATH = Path(__file__).parent.parent.parent.parent / "data" / "bot-prod-20260224-195317.db"
OUTPUT_PATH = Path(__file__).parent / "2026-02-24-usage-report.xlsx"
DEV_IDS = (110859240, 437936743, 7964380536)
COST_PER_MINUTE = 0.67  # RUB per audio minute (OpenAI Whisper API)
START_DATE = "2025-12-15"

# Duration categories (label, min_seconds, max_seconds)
DURATION_BINS = [
    ("0-30 sec", 0, 30),
    ("31-60 sec", 31, 60),
    ("1-2 min", 61, 120),
    ("2-5 min", 121, 300),
    ("5-10 min", 301, 600),
    ("10-30 min", 601, 1800),
    ("30-60 min", 1801, 3600),
    ("60+ min", 3601, 999999),
]

# ── Styling ─────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
NUMBER_FMT_PCT = "0.0%"

# Chart colors — consistent palette for series
CHART_BLUE = "4472C4"
CHART_ORANGE = "ED7D31"
CHART_GREEN = "70AD47"
CHART_RED = "FF4B4B"


def style_header(ws: Any, ncols: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def auto_width(ws: Any) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def write_df(ws: Any, df: pd.DataFrame, start_row: int = 1) -> int:
    """Write DataFrame to worksheet and return next empty row."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    last_row = start_row + len(df)
    return last_row + 1


def add_autofilter(ws: Any, ncols: int, nrows: int) -> None:
    """Add autofilter to the header row."""
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"


def set_series_color(chart: Any, series_idx: int, hex_color: str) -> None:
    """Set a solid fill color on a chart series so all bars/points are one color."""
    series = chart.series[series_idx]
    series.graphicalProperties.solidFill = hex_color


def categorize_duration(seconds: Any) -> str:
    """Assign a duration category label based on seconds."""
    if pd.isna(seconds) or seconds is None:
        return "Unknown"
    s = int(seconds)
    for label, lo, hi in DURATION_BINS:
        if lo <= s <= hi:
            return label
    return "60+ min"


# ── Data Loading ────────────────────────────────────────────────────────────
def load_data() -> dict[str, pd.DataFrame]:
    """Load all needed data from SQLite."""
    conn = sqlite3.connect(DB_PATH)

    dev_filter = f"u.telegram_id NOT IN {DEV_IDS}"

    # Users with aggregated usage
    users_df = pd.read_sql_query(
        f"""
        SELECT
            u.id as user_id,
            u.telegram_id,
            u.username,
            u.first_name,
            u.last_name,
            u.created_at as registered_at,
            u.total_usage_seconds,
            COUNT(us.id) as transcription_count,
            COALESCE(SUM(us.voice_duration_seconds), 0) as total_audio_seconds,
            COALESCE(SUM(us.voice_duration_seconds) / 60.0, 0) as total_audio_minutes,
            MIN(us.created_at) as first_usage,
            MAX(us.created_at) as last_usage,
            COUNT(DISTINCT DATE(us.created_at)) as active_days
        FROM users u
        LEFT JOIN usage us ON u.id = us.user_id
        WHERE {dev_filter}
          AND u.created_at >= '{START_DATE}'
        GROUP BY u.id
        ORDER BY total_audio_seconds DESC
        """,
        conn,
    )
    users_df["cost_rub"] = users_df["total_audio_minutes"] * COST_PER_MINUTE

    # Usage records
    usage_df = pd.read_sql_query(
        f"""
        SELECT
            us.id as usage_id,
            us.user_id,
            u.telegram_id,
            u.username,
            u.first_name,
            DATE(us.created_at) as date,
            us.created_at,
            us.voice_duration_seconds,
            ROUND(us.voice_duration_seconds / 60.0, 1) as voice_duration_minutes,
            us.model_size,
            ROUND(us.processing_time_seconds, 1) as processing_time,
            us.transcription_length,
            us.llm_model,
            us.llm_processing_time_seconds,
            strftime('%H', us.created_at) as hour_utc
        FROM usage us
        JOIN users u ON u.id = us.user_id
        WHERE {dev_filter}
          AND us.created_at >= '{START_DATE}'
        ORDER BY us.created_at
        """,
        conn,
    )

    # Transcription variants (button usage)
    variants_df = pd.read_sql_query(
        f"""
        SELECT
            tv.id,
            tv.usage_id,
            tv.mode,
            tv.length_level,
            tv.emoji_level,
            tv.timestamps_enabled,
            tv.generated_by,
            tv.llm_model,
            tv.processing_time_seconds,
            tv.created_at,
            us.user_id,
            u.username
        FROM transcription_variants tv
        JOIN usage us ON us.id = tv.usage_id
        JOIN users u ON u.id = us.user_id
        WHERE {dev_filter}
          AND tv.created_at >= '{START_DATE}'
        """,
        conn,
    )

    conn.close()
    return {"users": users_df, "usage": usage_df, "variants": variants_df}


# ── Sheet Builders ──────────────────────────────────────────────────────────


def build_users_sheet(wb: Workbook, users_df: pd.DataFrame) -> None:
    """Sheet 1: Users — all users with aggregated metrics."""
    ws = wb.create_sheet("Users")
    df = users_df[
        [
            "telegram_id",
            "username",
            "first_name",
            "registered_at",
            "total_usage_seconds",
            "transcription_count",
            "total_audio_minutes",
            "cost_rub",
            "active_days",
            "first_usage",
            "last_usage",
        ]
    ].copy()
    df.columns = [
        "Telegram ID",
        "Username",
        "First Name",
        "Registered",
        "Total Usage (sec)",
        "Transcriptions",
        "Audio Minutes",
        "Cost (RUB)",
        "Active Days",
        "First Usage",
        "Last Usage",
    ]

    nrows = write_df(ws, df)
    style_header(ws, len(df.columns))
    add_autofilter(ws, len(df.columns), nrows - 2)
    auto_width(ws)


def build_usage_sheet(wb: Workbook, usage_df: pd.DataFrame) -> None:
    """Sheet 2: Usage — all usage records."""
    ws = wb.create_sheet("Usage")
    df = usage_df[
        [
            "date",
            "telegram_id",
            "username",
            "first_name",
            "voice_duration_seconds",
            "voice_duration_minutes",
            "model_size",
            "processing_time",
            "transcription_length",
            "llm_model",
            "hour_utc",
        ]
    ].copy()
    df.columns = [
        "Date",
        "Telegram ID",
        "Username",
        "First Name",
        "Duration (sec)",
        "Duration (min)",
        "Model",
        "Processing Time (sec)",
        "Transcription Length",
        "LLM Model",
        "Hour (UTC)",
    ]

    nrows = write_df(ws, df)
    style_header(ws, len(df.columns))
    add_autofilter(ws, len(df.columns), nrows - 2)
    auto_width(ws)


def build_daily_summary(wb: Workbook, usage_df: pd.DataFrame, users_df: pd.DataFrame) -> None:
    """Sheet 3: Daily Summary with charts."""
    ws = wb.create_sheet("Daily Summary")

    # New registrations per day
    reg_df = users_df.copy()
    reg_df["reg_date"] = pd.to_datetime(reg_df["registered_at"]).dt.date.astype(str)
    new_regs = reg_df.groupby("reg_date").size().reset_index(name="new_registrations")

    daily = (
        usage_df.groupby("date")
        .agg(
            dau=("username", "nunique"),
            transcriptions=("usage_id", "count"),
            audio_minutes=("voice_duration_minutes", "sum"),
            base_fallbacks=("model_size", lambda x: (x == "base").sum()),
            null_events=("model_size", lambda x: x.isna().sum()),
        )
        .reset_index()
    )
    daily["cost_rub"] = daily["audio_minutes"] * COST_PER_MINUTE
    daily = daily.merge(new_regs, left_on="date", right_on="reg_date", how="left").drop(
        columns=["reg_date"], errors="ignore"
    )
    daily["new_registrations"] = daily["new_registrations"].fillna(0).astype(int)

    daily.columns = [
        "Date",
        "DAU",
        "Transcriptions",
        "Audio Minutes",
        "Base Fallbacks",
        "NULL Events",
        "Cost (RUB)",
        "New Registrations",
    ]

    write_df(ws, daily)
    n_data = len(daily)
    style_header(ws, len(daily.columns))
    auto_width(ws)

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)

    # Line chart: DAU
    chart1 = LineChart()
    chart1.title = "Daily Active Users (DAU)"
    chart1.y_axis.title = "Users"
    chart1.x_axis.title = "Date"
    chart1.width = 28
    chart1.height = 14
    data = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.style = 10
    set_series_color(chart1, 0, CHART_BLUE)
    ws.add_chart(chart1, f"A{n_data + 4}")

    # Bar chart: Cost
    chart2 = BarChart()
    chart2.title = "Daily Cost (RUB)"
    chart2.y_axis.title = "RUB"
    chart2.width = 28
    chart2.height = 14
    data2 = Reference(ws, min_col=7, min_row=1, max_row=n_data + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.style = 10
    set_series_color(chart2, 0, CHART_ORANGE)
    ws.add_chart(chart2, f"A{n_data + 20}")


def build_weekly_summary(wb: Workbook, usage_df: pd.DataFrame, users_df: pd.DataFrame) -> None:
    """Sheet 4: Weekly Summary with charts."""
    ws = wb.create_sheet("Weekly Summary")

    udf = usage_df.copy()
    udf["week"] = pd.to_datetime(udf["date"]).dt.strftime("%Y-W%V")

    # New registrations per week
    reg_df = users_df.copy()
    reg_df["reg_week"] = pd.to_datetime(reg_df["registered_at"]).dt.strftime("%Y-W%V")
    new_regs = reg_df.groupby("reg_week").size().reset_index(name="new_users")

    weekly = (
        udf.groupby("week")
        .agg(
            wau=("username", "nunique"),
            transcriptions=("usage_id", "count"),
            audio_minutes=("voice_duration_minutes", "sum"),
        )
        .reset_index()
    )
    weekly["cost_rub"] = weekly["audio_minutes"] * COST_PER_MINUTE
    weekly = weekly.merge(new_regs, left_on="week", right_on="reg_week", how="left").drop(
        columns=["reg_week"], errors="ignore"
    )
    weekly["new_users"] = weekly["new_users"].fillna(0).astype(int)

    weekly.columns = [
        "Week",
        "WAU",
        "Transcriptions",
        "Audio Minutes",
        "Cost (RUB)",
        "New Users",
    ]

    write_df(ws, weekly)
    n_data = len(weekly)
    style_header(ws, len(weekly.columns))
    auto_width(ws)

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)

    # Line chart: WAU
    chart1 = LineChart()
    chart1.title = "Weekly Active Users (WAU)"
    chart1.y_axis.title = "Users"
    chart1.width = 24
    chart1.height = 14
    data = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.style = 10
    set_series_color(chart1, 0, CHART_BLUE)
    ws.add_chart(chart1, f"A{n_data + 4}")

    # Bar chart: Cost
    chart2 = BarChart()
    chart2.title = "Weekly Cost (RUB)"
    chart2.y_axis.title = "RUB"
    chart2.width = 24
    chart2.height = 14
    data2 = Reference(ws, min_col=5, min_row=1, max_row=n_data + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.style = 10
    set_series_color(chart2, 0, CHART_ORANGE)
    ws.add_chart(chart2, f"A{n_data + 20}")


def build_monthly_summary(wb: Workbook, usage_df: pd.DataFrame) -> None:
    """Sheet 5: Monthly Summary with charts."""
    ws = wb.create_sheet("Monthly Summary")

    udf = usage_df.copy()
    udf["month"] = pd.to_datetime(udf["date"]).dt.strftime("%Y-%m")

    monthly = (
        udf.groupby("month")
        .agg(
            mau=("username", "nunique"),
            transcriptions=("usage_id", "count"),
            audio_minutes=("voice_duration_minutes", "sum"),
        )
        .reset_index()
    )
    monthly["cost_rub"] = monthly["audio_minutes"] * COST_PER_MINUTE

    monthly.columns = ["Month", "MAU", "Transcriptions", "Audio Minutes", "Cost (RUB)"]

    write_df(ws, monthly)
    n_data = len(monthly)
    style_header(ws, len(monthly.columns))
    auto_width(ws)

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)

    # Bar chart: MAU
    chart1 = BarChart()
    chart1.title = "Monthly Active Users (MAU)"
    chart1.y_axis.title = "MAU"
    chart1.width = 20
    chart1.height = 14
    data = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.style = 10
    set_series_color(chart1, 0, CHART_BLUE)
    ws.add_chart(chart1, f"A{n_data + 4}")


def build_model_distribution(wb: Workbook, usage_df: pd.DataFrame) -> None:
    """Sheet 6: Model Distribution with pie chart."""
    ws = wb.create_sheet("Model Distribution")

    udf = usage_df.copy()
    udf["model_size"] = udf["model_size"].fillna("NULL (failed)")

    model_stats = (
        udf.groupby("model_size")
        .agg(
            count=("usage_id", "count"),
            avg_processing_time=("processing_time", "mean"),
            total_audio_min=("voice_duration_minutes", "sum"),
        )
        .reset_index()
    )
    total = model_stats["count"].sum()
    model_stats["pct"] = model_stats["count"] / total
    model_stats["avg_processing_time"] = model_stats["avg_processing_time"].round(1)
    model_stats["total_audio_min"] = model_stats["total_audio_min"].round(1)
    model_stats = model_stats.sort_values("count", ascending=False)

    model_stats.columns = [
        "Model",
        "Count",
        "Avg Processing Time (sec)",
        "Total Audio (min)",
        "% of Total",
    ]

    write_df(ws, model_stats)
    n_data = len(model_stats)
    style_header(ws, len(model_stats.columns))
    auto_width(ws)

    # Format percentage column
    for row in range(2, n_data + 2):
        ws.cell(row=row, column=5).number_format = NUMBER_FMT_PCT

    # Pie chart
    chart = PieChart()
    chart.title = "Model Distribution"
    chart.width = 20
    chart.height = 14
    data = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    ws.add_chart(chart, f"A{n_data + 4}")


def build_button_usage(wb: Workbook, variants_df: pd.DataFrame) -> None:
    """Sheet 7: Button Usage (transcription variant modes)."""
    ws = wb.create_sheet("Button Usage")

    all_modes = (
        variants_df.groupby("mode")
        .agg(
            count=("id", "count"),
            unique_users=("user_id", "nunique"),
        )
        .reset_index()
    )
    total = all_modes["count"].sum()
    all_modes["pct"] = all_modes["count"] / total
    all_modes = all_modes.sort_values("count", ascending=False)
    all_modes.columns = ["Mode", "Variants Created", "Unique Users", "% of Total"]

    write_df(ws, all_modes)
    n_data = len(all_modes)
    style_header(ws, len(all_modes.columns))
    auto_width(ws)

    for row in range(2, n_data + 2):
        ws.cell(row=row, column=4).number_format = NUMBER_FMT_PCT

    # Bar chart
    chart = BarChart()
    chart.title = "Button Usage by Mode"
    chart.y_axis.title = "Count"
    chart.width = 20
    chart.height = 14
    data = Reference(ws, min_col=2, min_row=1, max_row=n_data + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 10
    set_series_color(chart, 0, CHART_GREEN)
    ws.add_chart(chart, f"A{n_data + 4}")


def build_hourly_activity(wb: Workbook, usage_df: pd.DataFrame) -> None:
    """Sheet 8: Hourly Activity (UTC and MSK)."""
    ws = wb.create_sheet("Hourly Activity")

    udf = usage_df.copy()
    udf["hour_utc"] = udf["hour_utc"].astype(int)

    hourly = (
        udf.groupby("hour_utc")
        .agg(
            transcriptions=("usage_id", "count"),
            unique_users=("username", "nunique"),
            audio_minutes=("voice_duration_minutes", "sum"),
        )
        .reset_index()
    )
    hourly["hour_msk"] = (hourly["hour_utc"] + 3) % 24
    hourly["audio_minutes"] = hourly["audio_minutes"].round(1)
    hourly = hourly.sort_values("hour_utc")

    hourly.columns = [
        "Hour (UTC)",
        "Transcriptions",
        "Unique Users",
        "Audio Minutes",
        "Hour (MSK)",
    ]
    # Reorder columns
    hourly = hourly[["Hour (UTC)", "Hour (MSK)", "Transcriptions", "Unique Users", "Audio Minutes"]]

    write_df(ws, hourly)
    n_data = len(hourly)
    style_header(ws, len(hourly.columns))
    auto_width(ws)

    # Bar chart
    chart = BarChart()
    chart.title = "Transcriptions by Hour (UTC)"
    chart.y_axis.title = "Transcriptions"
    chart.x_axis.title = "Hour (UTC)"
    chart.width = 28
    chart.height = 14
    data = Reference(ws, min_col=3, min_row=1, max_row=n_data + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_data + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.style = 10
    set_series_color(chart, 0, CHART_BLUE)
    ws.add_chart(chart, f"A{n_data + 4}")


def build_user_segments(wb: Workbook, users_df: pd.DataFrame, usage_df: pd.DataFrame) -> None:
    """Sheet 9: User Segments — recency-aware activity, cost, duration profile, negative UX."""
    ws = wb.create_sheet("User Segments")

    df = users_df[users_df["transcription_count"] > 0].copy()

    # ── Per-user duration distribution ──────────────────────────────────────
    udf = usage_df.copy()
    udf["dur_category"] = udf["voice_duration_seconds"].apply(categorize_duration)
    cat_labels = [label for label, _, _ in DURATION_BINS]

    pivot = udf.groupby(["user_id", "dur_category"]).size().unstack(fill_value=0)
    ordered_cols = [c for c in cat_labels if c in pivot.columns]
    pivot = pivot[ordered_cols].reset_index()
    df = df.merge(pivot, on="user_id", how="left")
    for c in ordered_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = df[c].fillna(0).astype(int)

    # ── Negative UX flag ────────────────────────────────────────────────────
    # Negative UX = base fallback OR NULL model OR file >7 min before bug fix
    TRUNCATION_FIX_DATE = "2026-02-23"
    udf["created_dt"] = pd.to_datetime(udf["created_at"])

    neg_base = udf[udf["model_size"] == "base"].groupby("user_id").size()
    neg_null = udf[udf["model_size"].isna()].groupby("user_id").size()
    neg_truncated = (
        udf[
            (udf["voice_duration_seconds"] > 420)
            & (udf["created_dt"] < TRUNCATION_FIX_DATE)
        ]
        .groupby("user_id")
        .size()
    )

    df["base_events"] = df["user_id"].map(neg_base).fillna(0).astype(int)
    df["null_events"] = df["user_id"].map(neg_null).fillna(0).astype(int)
    df["truncated_events"] = df["user_id"].map(neg_truncated).fillna(0).astype(int)
    df["negative_ux_events"] = df["base_events"] + df["null_events"] + df["truncated_events"]
    df["had_negative_ux"] = df["negative_ux_events"] > 0

    # Compute days since registration and days since last usage
    report_date = pd.Timestamp("2026-02-23")
    df["registered_dt"] = pd.to_datetime(df["registered_at"])
    df["last_usage_dt"] = pd.to_datetime(df["last_usage"])
    df["days_since_reg"] = (report_date - df["registered_dt"]).dt.days
    df["days_since_last"] = (report_date - df["last_usage_dt"]).dt.days

    # Activity segments (based on active_days)
    def activity_segment(row: pd.Series) -> str:
        days = row["active_days"]
        if days >= 30:
            return "Power User (30+ days)"
        elif days >= 15:
            return "Regular (15-29 days)"
        elif days >= 8:
            return "Moderate (8-14 days)"
        elif days >= 4:
            return "Light (4-7 days)"
        elif days >= 2:
            return "Casual (2-3 days)"
        else:
            return "One-time (1 day)"

    # Recency status: distinguishes "new user" from "churned"
    def recency_status(row: pd.Series) -> str:
        days_since_reg = row["days_since_reg"]
        days_since_last = row["days_since_last"]
        active_days = row["active_days"]

        if days_since_reg <= 7:
            return "New (<7d ago)"
        if days_since_last <= 7:
            return "Active (seen <7d)"
        if days_since_last <= 14:
            return "Cooling (8-14d)"
        if active_days == 1:
            return "Churned one-timer"
        return "Churned returning"

    # Cost segments
    def cost_segment(row: pd.Series) -> str:
        cost = row["cost_rub"]
        if cost >= 100:
            return "High (100+ RUB)"
        elif cost >= 20:
            return "Medium (20-99 RUB)"
        elif cost >= 5:
            return "Low (5-19 RUB)"
        else:
            return "Minimal (<5 RUB)"

    df["activity_segment"] = df.apply(activity_segment, axis=1)
    df["recency_status"] = df.apply(recency_status, axis=1)
    df["cost_segment"] = df.apply(cost_segment, axis=1)

    base_cols = [
        "telegram_id",
        "username",
        "first_name",
        "transcription_count",
        "total_audio_minutes",
        "cost_rub",
        "active_days",
        "activity_segment",
        "recency_status",
        "cost_segment",
        "days_since_reg",
        "days_since_last",
        "first_usage",
        "last_usage",
    ]
    dur_cols = ordered_cols  # duration category columns
    ux_cols = [
        "base_events",
        "null_events",
        "truncated_events",
        "negative_ux_events",
        "had_negative_ux",
    ]
    out = df[base_cols + dur_cols + ux_cols].copy()
    out.columns = (
        [
            "Telegram ID",
            "Username",
            "First Name",
            "Transcriptions",
            "Audio Minutes",
            "Cost (RUB)",
            "Active Days",
            "Activity Segment",
            "Recency Status",
            "Cost Segment",
            "Days Since Reg",
            "Days Since Last Use",
            "First Usage",
            "Last Usage",
        ]
        + dur_cols
        + [
            "Base Events",
            "NULL Events",
            "Truncated (>7min pre-fix)",
            "Total Negative UX Events",
            "Had Negative UX",
        ]
    )
    out = out.sort_values("Cost (RUB)", ascending=False)

    nrows = write_df(ws, out)
    style_header(ws, len(out.columns))
    add_autofilter(ws, len(out.columns), nrows - 2)
    auto_width(ws)


def build_negative_ux(wb: Workbook, usage_df: pd.DataFrame, users_df: pd.DataFrame) -> None:
    """Sheet 10: Negative UX — users who experienced base/NULL fallbacks."""
    ws = wb.create_sheet("Negative UX")

    udf = usage_df.copy()

    # Build user_id -> telegram_id lookup
    tid_map = users_df.set_index("user_id")["telegram_id"].to_dict()

    # Find users with base or NULL model events
    base_uids = udf[udf["model_size"] == "base"]["user_id"].unique()
    null_uids = udf[udf["model_size"].isna()]["user_id"].unique()
    affected_uids = set(base_uids) | set(null_uids)

    rows = []
    for uid in affected_uids:
        user_usage = udf[udf["user_id"] == uid].sort_values("created_at")
        total = len(user_usage)
        base_count = (user_usage["model_size"] == "base").sum()
        null_count = user_usage["model_size"].isna().sum()
        good_count = total - base_count - null_count

        negative_dates = user_usage[
            (user_usage["model_size"] == "base") | (user_usage["model_size"].isna())
        ]["created_at"]
        last_negative = negative_dates.max()
        good_after = user_usage[
            (user_usage["created_at"] > last_negative)
            & (user_usage["model_size"].notna())
            & (user_usage["model_size"] != "base")
        ]
        returned = len(good_after) > 0

        first_event_model = user_usage.iloc[0]["model_size"] if len(user_usage) > 0 else None
        first_was_negative = first_event_model == "base" or pd.isna(first_event_model)

        username = user_usage.iloc[0]["username"] if len(user_usage) > 0 else None

        rows.append(
            {
                "Telegram ID": tid_map.get(uid, uid),
                "Username": username,
                "Total Transcriptions": total,
                "Base Events": base_count,
                "NULL Events": null_count,
                "Good Events": good_count,
                "First Event Negative": first_was_negative,
                "Returned After Negative": returned,
                "First Usage": user_usage["date"].min(),
                "Last Usage": user_usage["date"].max(),
            }
        )

    neg_df = pd.DataFrame(rows).sort_values("Base Events", ascending=False)

    nrows = write_df(ws, neg_df)
    style_header(ws, len(neg_df.columns))
    add_autofilter(ws, len(neg_df.columns), nrows - 2)
    auto_width(ws)


def build_cohort_retention(wb: Workbook, usage_df: pd.DataFrame, users_df: pd.DataFrame) -> None:
    """Sheet 11: Cohort Retention — weekly cohorts with week-over-week retention."""
    ws = wb.create_sheet("Cohort Retention")

    udf = usage_df.copy()
    udf["usage_week"] = pd.to_datetime(udf["date"]).dt.strftime("%Y-W%V")

    # Map user_id -> registration week
    reg = users_df[["user_id", "registered_at"]].copy()
    reg["reg_week"] = pd.to_datetime(reg["registered_at"]).dt.strftime("%Y-W%V")

    # Get unique weeks each user was active
    user_weeks = udf.groupby("user_id")["usage_week"].apply(set).reset_index()
    user_weeks.columns = ["user_id", "active_weeks"]
    reg = reg.merge(user_weeks, on="user_id", how="left")

    # Build sorted list of all weeks
    all_weeks = sorted(set(reg["reg_week"].dropna()) | set(udf["usage_week"].dropna()))
    week_index = {w: i for i, w in enumerate(all_weeks)}

    max_rel_weeks = 10
    cohort_rows = []

    for cohort_week in all_weeks:
        cohort_users = reg[reg["reg_week"] == cohort_week]
        cohort_size = len(cohort_users)
        if cohort_size == 0:
            continue

        cohort_idx = week_index[cohort_week]
        row: dict[str, Any] = {"Cohort Week": cohort_week, "Cohort Size": cohort_size}

        for rel_w in range(0, max_rel_weeks + 1):
            target_idx = cohort_idx + rel_w
            if target_idx >= len(all_weeks):
                row[f"W+{rel_w}"] = None
                continue
            target_week = all_weeks[target_idx]
            active_count = 0
            for _, user_row in cohort_users.iterrows():
                active_set = user_row["active_weeks"]
                if isinstance(active_set, set) and target_week in active_set:
                    active_count += 1
            row[f"W+{rel_w}"] = active_count / cohort_size if cohort_size > 0 else 0

        cohort_rows.append(row)

    cohort_df = pd.DataFrame(cohort_rows)
    write_df(ws, cohort_df)
    n_data = len(cohort_df)
    style_header(ws, len(cohort_df.columns))
    auto_width(ws)

    # Format retention percentages
    for row_idx in range(2, n_data + 2):
        for col_idx in range(3, len(cohort_df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = NUMBER_FMT_PCT

    # Line chart for top cohorts
    if n_data >= 2:
        chart = LineChart()
        chart.title = "Cohort Retention by Week"
        chart.y_axis.title = "Retention %"
        chart.x_axis.title = "Relative Week"
        chart.width = 28
        chart.height = 14
        cats = Reference(ws, min_col=3, max_col=min(13, len(cohort_df.columns)), min_row=1)
        chart.set_categories(cats)
        top_cohorts = cohort_df.nlargest(6, "Cohort Size")
        for _, c_row in top_cohorts.iterrows():
            row_num = int(cohort_df.index.get_loc(c_row.name)) + 2
            data = Reference(
                ws,
                min_col=3,
                max_col=min(13, len(cohort_df.columns)),
                min_row=row_num,
                max_row=row_num,
            )
            chart.add_data(data, from_rows=True)
            chart.series[-1].tx = SeriesLabel(v=c_row["Cohort Week"])
        ws.add_chart(chart, f"A{n_data + 4}")


def build_duration_distribution(wb: Workbook, usage_df: pd.DataFrame, users_df: pd.DataFrame) -> None:
    """Sheet 12: Duration Distribution — overall summary + per-user file length breakdown."""
    # ── Part 1: Overall summary ─────────────────────────────────────────────
    ws = wb.create_sheet("Duration Distribution")

    udf = usage_df.copy()
    udf["dur_category"] = udf["voice_duration_seconds"].apply(categorize_duration)

    # Ordered categories
    cat_labels = [label for label, _, _ in DURATION_BINS]
    cat_order = {label: i for i, label in enumerate(cat_labels)}

    summary_rows = []
    total_files = len(udf)
    for label, lo, hi in DURATION_BINS:
        subset = udf[udf["dur_category"] == label]
        count = len(subset)
        pct = count / total_files if total_files > 0 else 0
        audio_min = subset["voice_duration_minutes"].sum()
        cost = audio_min * COST_PER_MINUTE
        unique_users = subset["user_id"].nunique()
        avg_duration = subset["voice_duration_seconds"].mean() if count > 0 else 0
        summary_rows.append(
            {
                "Category": label,
                "Files": count,
                "% of Files": pct,
                "Unique Users": unique_users,
                "Audio Minutes": round(audio_min, 1),
                "Cost (RUB)": round(cost, 1),
                "% of Cost": 0,  # filled below
                "Avg Duration (sec)": round(avg_duration, 1),
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    total_cost = summary_df["Cost (RUB)"].sum()
    if total_cost > 0:
        summary_df["% of Cost"] = summary_df["Cost (RUB)"] / total_cost

    write_df(ws, summary_df)
    n_sum = len(summary_df)
    style_header(ws, len(summary_df.columns))
    auto_width(ws)

    # Format percentage columns
    for row_idx in range(2, n_sum + 2):
        ws.cell(row=row_idx, column=3).number_format = NUMBER_FMT_PCT  # % of Files
        ws.cell(row=row_idx, column=7).number_format = NUMBER_FMT_PCT  # % of Cost

    cats = Reference(ws, min_col=1, min_row=2, max_row=n_sum + 1)

    # Bar chart: File count by category
    chart1 = BarChart()
    chart1.title = "Files by Duration Category"
    chart1.y_axis.title = "Files"
    chart1.width = 24
    chart1.height = 14
    data1 = Reference(ws, min_col=2, min_row=1, max_row=n_sum + 1)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.style = 10
    set_series_color(chart1, 0, CHART_BLUE)
    ws.add_chart(chart1, f"A{n_sum + 4}")

    # Bar chart: Cost by category
    chart2 = BarChart()
    chart2.title = "Cost by Duration Category (RUB)"
    chart2.y_axis.title = "RUB"
    chart2.width = 24
    chart2.height = 14
    data2 = Reference(ws, min_col=6, min_row=1, max_row=n_sum + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.style = 10
    set_series_color(chart2, 0, CHART_ORANGE)
    ws.add_chart(chart2, f"A{n_sum + 20}")

    # ── Part 2: Per-user duration breakdown ─────────────────────────────────
    ws2 = wb.create_sheet("User Duration Profile")

    # Pivot: user_id -> count per category
    udf["cat_sort"] = udf["dur_category"].map(cat_order)
    pivot = udf.groupby(["user_id", "dur_category"]).size().unstack(fill_value=0)
    # Reorder columns by duration
    ordered_cols = [c for c in cat_labels if c in pivot.columns]
    pivot = pivot[ordered_cols]

    # Join user info
    user_info = users_df[["user_id", "telegram_id", "username", "first_name"]].copy()
    pivot = pivot.reset_index().merge(user_info, on="user_id", how="left")

    # Compute totals and percentages for the "long files" share
    long_cats = ["10-30 min", "30-60 min", "60+ min"]
    pivot["total_files"] = pivot[ordered_cols].sum(axis=1)
    long_cols_present = [c for c in long_cats if c in pivot.columns]
    pivot["long_files"] = pivot[long_cols_present].sum(axis=1) if long_cols_present else 0
    pivot["long_pct"] = pivot["long_files"] / pivot["total_files"]

    # Reorder columns for output
    out_cols = ["telegram_id", "username", "first_name", "total_files"] + ordered_cols + [
        "long_files",
        "long_pct",
    ]
    out = pivot[out_cols].sort_values("total_files", ascending=False)
    out.columns = (
        ["Telegram ID", "Username", "First Name", "Total Files"]
        + ordered_cols
        + ["Long Files (10+ min)", "Long %"]
    )

    nrows2 = write_df(ws2, out)
    style_header(ws2, len(out.columns))
    add_autofilter(ws2, len(out.columns), nrows2 - 2)
    auto_width(ws2)

    # Format long_pct column
    long_pct_col = len(out.columns)  # last column
    for row_idx in range(2, nrows2):
        ws2.cell(row=row_idx, column=long_pct_col).number_format = NUMBER_FMT_PCT


# ── Main ────────────────────────────────────────────────────────────────────
def main() -> None:
    print(f"Loading data from {DB_PATH}...")
    data = load_data()
    print(
        f"  Users: {len(data['users'])}, Usage: {len(data['usage'])}, "
        f"Variants: {len(data['variants'])}"
    )

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    print("Building sheets...")
    build_users_sheet(wb, data["users"])
    print("  ✓ Users")
    build_usage_sheet(wb, data["usage"])
    print("  ✓ Usage")
    build_daily_summary(wb, data["usage"], data["users"])
    print("  ✓ Daily Summary")
    build_weekly_summary(wb, data["usage"], data["users"])
    print("  ✓ Weekly Summary")
    build_monthly_summary(wb, data["usage"])
    print("  ✓ Monthly Summary")
    build_model_distribution(wb, data["usage"])
    print("  ✓ Model Distribution")
    build_button_usage(wb, data["variants"])
    print("  ✓ Button Usage")
    build_hourly_activity(wb, data["usage"])
    print("  ✓ Hourly Activity")
    build_user_segments(wb, data["users"], data["usage"])
    print("  ✓ User Segments")
    build_negative_ux(wb, data["usage"], data["users"])
    print("  ✓ Negative UX")
    build_cohort_retention(wb, data["usage"], data["users"])
    print("  ✓ Cohort Retention")
    build_duration_distribution(wb, data["usage"], data["users"])
    print("  ✓ Duration Distribution")
    print("  ✓ User Duration Profile")

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(str(OUTPUT_PATH))
    print("Done!")


if __name__ == "__main__":
    main()
